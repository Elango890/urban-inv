# =============================================================================
# reports/views.py  —  Production-Grade Report Module
#
# Endpoints:
#   GET  /api/reports/types/    → Report category + report list
#   GET  /api/reports/stats/    → Dashboard summary statistics
#   POST /api/reports/preview/  → Paginated preview (JSON)
#   POST /api/reports/generate/ → File download (PDF / XLSX / CSV)
#
# All builders return:  { "title": str, "headers": [...], "rows": [[...]] }
#
# Model map (verified against uploaded model files):
#   masters   : Asset (asset_code/name/category/purchase_cost/selling_price/status)
#               Supplier (display_name/company_name/first_name/last_name/phone/mobile/
#                         trn/gstin/outstanding/credit_limit/is_active)
#               Customer (display_name/company_name/first_name/last_name/phone/mobile/
#                         trn/email/outstanding/credit_limit/is_active/customer_type)
#               FinancialYear (year_name/start_date/end_date/is_active)
#   sales     : SalesInvoice / SalesInvoiceItem / SalesPayment
#   purchases : PurchaseEntry / PurchaseOrder / PurchasePayment
#   stock     : Stock / StockHistory / Warehouse
# =============================================================================

from __future__ import annotations

import csv
import io
import logging
import os
import traceback
from datetime import date as _date, datetime
from decimal import Decimal

from django.db.models import Count, F, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from users.permissions import HasAllowedRoles, SALES_ROLES

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Lazy model loader — avoids circular imports at module level
# ─────────────────────────────────────────────────────────────────────────────

def _M() -> dict:
    """Return all required model classes as a dict."""
    from masters.models import Asset, Customer, FinancialYear, Supplier
    from purchases.models import PurchaseEntry, PurchaseOrder, PurchasePayment
    from sales.models import (
        SalesInvoice,
        SalesInvoiceItem,
        SalesPayment,
        SalesReturn,
        SalesReturnItem,
    )
    from stock.models import Stock, StockHistory, Warehouse

    return {
        "Asset": Asset,
        "Supplier": Supplier,
        "Customer": Customer,
        "FinancialYear": FinancialYear,
        "SalesInvoice": SalesInvoice,
        "SalesInvoiceItem": SalesInvoiceItem,
        "SalesPayment": SalesPayment,
        "SalesReturn": SalesReturn,
        "SalesReturnItem": SalesReturnItem,
        "PurchaseEntry": PurchaseEntry,
        "PurchaseOrder": PurchaseOrder,
        "PurchasePayment": PurchasePayment,
        "Stock": Stock,
        "StockHistory": StockHistory,
        "Warehouse": Warehouse,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Formatting helpers
# ─────────────────────────────────────────────────────────────────────────────

def _f(v) -> float:
    """Safe float conversion."""
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _aed(v) -> str:
    """Format a number as AED currency string."""
    n = _f(v)
    abs_n = abs(n)
    sign = "-" if n < 0 else ""
    if abs_n >= 1_000_000:
        return f"{sign}AED {abs_n / 1_000_000:.2f}M"
    if abs_n >= 1_000:
        return f"{sign}AED {abs_n / 1_000:.2f}K"
    return f"{sign}AED {abs_n:,.2f}"


def _num(v) -> str:
    return f"{_f(v):,.2f}"


def _pct(v) -> str:
    return f"{_f(v):.1f}%"


def _ds(d) -> str:
    """Format date."""
    if d is None:
        return "—"
    try:
        return d.strftime("%d %b %Y")
    except AttributeError:
        return str(d)


def _dts(d) -> str:
    """Format datetime."""
    if d is None:
        return "—"
    try:
        return d.strftime("%d %b %Y %H:%M")
    except AttributeError:
        return str(d)[:16]


def _str(v) -> str:
    """Safe string, returns em-dash for falsy values."""
    return str(v).strip() if v else "—"


def _upper(v) -> str:
    return (v or "").upper() or "—"


def _title(v) -> str:
    return (v or "").replace("_", " ").title() or "—"


def _sales_invoice_financials(invoice) -> dict[str, float]:
    gross_total = Decimal(str(invoice.total_amount or 0))
    raw_paid = Decimal(str(invoice.paid_amount or 0))
    return_total = Decimal(
        str(
            invoice.returns.filter(status="confirmed").aggregate(t=Sum("total_amount"))["t"]
            or 0
        )
    )
    net_total = max(gross_total - return_total, Decimal("0"))
    applied_paid = min(raw_paid, net_total)
    refundable_amount = max(raw_paid - net_total, Decimal("0"))
    balance_amount = max(net_total - raw_paid, Decimal("0"))

    if net_total <= Decimal("0.01"):
        payment_status = "paid"
    elif applied_paid <= Decimal("0.00"):
        payment_status = "unpaid"
    elif balance_amount <= Decimal("0.01"):
        payment_status = "paid"
    else:
        payment_status = "partial"

    return {
        "net_total": float(net_total),
        "paid_amount": float(applied_paid),
        "refundable_amount": float(refundable_amount),
        "balance_amount": float(balance_amount),
        "payment_status": payment_status,
        "payment_status_display": payment_status.replace("_", " ").title(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Name helpers — handles all Supplier / Customer name variants
# ─────────────────────────────────────────────────────────────────────────────

def _supplier_name(supplier) -> str:
    if not supplier:
        return "—"
    for attr in ("display_name", "company_name"):
        val = getattr(supplier, attr, "")
        if val:
            return val
    # Compose from parts
    parts = []
    for attr in ("first_name", "last_name"):
        v = getattr(supplier, attr, "")
        if v:
            parts.append(v)
    return " ".join(parts) or "—"


def _customer_name(customer) -> str:
    if not customer:
        return "—"
    for attr in ("display_name", "company_name"):
        val = getattr(customer, attr, "")
        if val:
            return val
    parts = []
    for attr in ("first_name", "last_name"):
        v = getattr(customer, attr, "")
        if v:
            parts.append(v)
    return " ".join(parts) or "—"


# ─────────────────────────────────────────────────────────────────────────────
# Q-object builders for fuzzy search
# ─────────────────────────────────────────────────────────────────────────────

def _supplier_fk_q(term: str) -> Q:
    """Q for filtering through a supplier FK (e.g. on PurchaseEntry)."""
    return (
        Q(vendor__display_name__icontains=term)
        | Q(vendor__company_name__icontains=term)
        | Q(vendor__first_name__icontains=term)
        | Q(vendor__last_name__icontains=term)
        | Q(vendor__phone__icontains=term)
        | Q(vendor__mobile__icontains=term)
    )


def _supplier_obj_q(term: str) -> Q:
    """Q for filtering Supplier objects directly."""
    return (
        Q(display_name__icontains=term)
        | Q(company_name__icontains=term)
        | Q(first_name__icontains=term)
        | Q(last_name__icontains=term)
        | Q(phone__icontains=term)
        | Q(mobile__icontains=term)
    )


def _customer_obj_q(term: str) -> Q:
    """Q for filtering Customer objects directly."""
    return (
        Q(display_name__icontains=term)
        | Q(company_name__icontains=term)
        | Q(first_name__icontains=term)
        | Q(last_name__icontains=term)
        | Q(phone__icontains=term)
        | Q(mobile__icontains=term)
    )


# ─────────────────────────────────────────────────────────────────────────────
# Financial Year & date helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_fy(fy_id):
    """Fetch a FinancialYear; falls back to the active one."""
    try:
        from masters.models import FinancialYear

        if fy_id:
            fy = FinancialYear.objects.filter(pk=fy_id).first()
            if fy:
                return fy
        return FinancialYear.objects.filter(is_active=True).first()
    except Exception:
        return None


def _bounds(fy_id, date_from, date_to):
    """
    Resolve the effective date range.

    Priority: explicit dates > FY bounds.
    Returns (f, t, fy) where f/t are strings or None.
    """
    fy = _get_fy(fy_id)
    f = date_from or (str(fy.start_date) if fy else None)
    t = date_to or (str(fy.end_date) if fy else None)
    return f, t, fy


def _fy_id(data: dict):
    return data.get("financialYearId") or data.get("financial_year")


def _report_company():
    from django.conf import settings
    from masters.models import OrganizationAddress

    company = {
        "name": "Urban Health Food Supplements Trading LLC",
        "trn": getattr(settings, "SALES_INVOICE_TRN", "") or "",
        "address_lines": [],
        "city_state_zip": "",
        "country": "United Arab Emirates",
        "phone": "",
        "title": "REPORT",
    }

    company.update(getattr(settings, "SALES_INVOICE_COMPANY", {}) or {})

    org = (
        OrganizationAddress.objects.filter(is_active=True, is_default=True).first()
        or OrganizationAddress.objects.filter(is_active=True).first()
    )
    if org:
        company["name"] = org.name or company["name"]
        company["address_lines"] = [
            line
            for line in [org.address_line1.strip(), org.address_line2.strip()]
            if line
        ]
        company["city_state_zip"] = ", ".join(
            part for part in [org.city.strip(), org.state.strip(), org.zip.strip()] if part
        )
        company["country"] = org.country or company["country"]
        company["phone"] = org.phone or company["phone"]

    return company


def _report_logo_path():
    from django.conf import settings

    configured = getattr(settings, "SALES_INVOICE_LOGO_PATH", "")
    candidates = [
        configured,
        os.path.join(settings.BASE_DIR, "inventory_backend", "static", "images", "logo.jpeg"),
        os.path.join(settings.BASE_DIR, "inventory_backend", "static", "images", "logo.png"),
    ]
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Internal exception
# ─────────────────────────────────────────────────────────────────────────────

class ReportBuildError(Exception):
    pass


# =============================================================================
# GET /api/reports/types/
# =============================================================================

class ReportTypesView(APIView):
    permission_classes = [IsAuthenticated, HasAllowedRoles]
    allowed_roles = SALES_ROLES

    def get(self, request):
        return Response(
            {
                "categories": [
                    {
                        "title": "Sales Reports",
                        "icon": "TrendingUp",
                        "color": "bg-emerald-500/10 text-emerald-600",
                        "reports": [
                            {
                                "id": "sales-summary",
                                "name": "Sales Summary",
                                "description": "All invoices with amounts, VAT, collected and balance",
                            },
                            {
                                "id": "monthly-sales",
                                "name": "Monthly Sales",
                                "description": "Month-by-month revenue with VAT breakdown",
                            },
                            {
                                "id": "daily-sales",
                                "name": "Daily Sales",
                                "description": "Day-wise invoice count and revenue",
                            },
                            {
                                "id": "sales-by-customer",
                                "name": "Sales by Customer",
                                "description": "Revenue and outstanding grouped by customer",
                            },
                            {
                                "id": "sales-by-item",
                                "name": "Sales by Item",
                                "description": "Revenue per asset / service line item",
                            },
                            {
                                "id": "sales-by-salesperson",
                                "name": "Sales by Salesperson",
                                "description": "Invoice count, revenue, collections and balance by salesperson",
                            },
                            {
                                "id": "salesperson-collections",
                                "name": "Salesperson Collections",
                                "description": "Payments received mapped back to the invoice salesperson",
                            },
                            {
                                "id": "sales-returns",
                                "name": "Sales Returns",
                                "description": "Return documents with customer, invoice, warehouse, amount and stock posting status",
                            },
                            {
                                "id": "sales-returns-by-customer",
                                "name": "Sales Returns by Customer",
                                "description": "Return count and credit value grouped by customer",
                            },
                            {
                                "id": "payment-collection",
                                "name": "Payment Collection",
                                "description": "All payments received with method and reference",
                            },
                            {
                                "id": "outstanding",
                                "name": "Outstanding Dues",
                                "description": "Unpaid & partially paid invoices with overdue days",
                            },
                            {
                                "id": "profit-loss",
                                "name": "Profit & Loss",
                                "description": "Sales revenue vs purchase cost, gross margin",
                            },
                        ],
                    },
                    {
                        "title": "Purchase Reports",
                        "icon": "Package",
                        "color": "bg-sky-500/10 text-sky-600",
                        "reports": [
                            {
                                "id": "purchase-summary",
                                "name": "Purchase Summary",
                                "description": "All purchase entries with payment status",
                            },
                            {
                                "id": "purchase-by-supplier",
                                "name": "Purchase by Supplier",
                                "description": "Spend, payments and outstanding per supplier",
                            },
                            {
                                "id": "purchase-entries",
                                "name": "Purchase Entries (GRN)",
                                "description": "Goods receipt notes — received and pending",
                            },
                            {
                                "id": "purchase-orders",
                                "name": "Purchase Orders",
                                "description": "PO list with status and amounts",
                            },
                            {
                                "id": "purchase-payments",
                                "name": "Purchase Payments",
                                "description": "All payments made to suppliers",
                            },
                            {
                                "id": "supplier-outstanding",
                                "name": "Supplier Outstanding",
                                "description": "Unpaid balances per active supplier",
                            },
                        ],
                    },
                    {
                        "title": "Inventory Reports",
                        "icon": "Warehouse",
                        "color": "bg-indigo-500/10 text-indigo-600",
                        "reports": [
                            {
                                "id": "stock-summary",
                                "name": "Stock Summary",
                                "description": "Current stock levels by asset and warehouse",
                            },
                            {
                                "id": "stock-low",
                                "name": "Low / Out-of-Stock",
                                "description": "Assets at or below minimum threshold",
                            },
                            {
                                "id": "stock-movement",
                                "name": "Stock Movement Log",
                                "description": "All in / out movements with running balance",
                            },
                        ],
                    },
                    {
                        "title": "Supplier & Customer",
                        "icon": "Users",
                        "color": "bg-violet-500/10 text-violet-600",
                        "reports": [
                            {
                                "id": "supplier-list",
                                "name": "Supplier Directory",
                                "description": "All suppliers with outstanding balances",
                            },
                            {
                                "id": "supplier-detail",
                                "name": "Supplier Detail",
                                "description": "Full supplier profile with linked purchase orders, entries and payments",
                            },
                            {
                                "id": "supplier-statement",
                                "name": "Supplier Statement",
                                "description": "Running supplier statement with purchase entries, payments and petty cash activity for a selected period",
                            },
                            {
                                "id": "customer-list",
                                "name": "Customer Directory",
                                "description": "All customers with outstanding balances",
                            },
                            {
                                "id": "customer-detail",
                                "name": "Customer Detail",
                                "description": "Full customer profile with linked invoices, payments and sales returns",
                            },
                            {
                                "id": "customer-statement",
                                "name": "Customer Statement",
                                "description": "Running customer statement with invoice, payment and return activity for a selected period",
                            },
                        ],
                    },
                ]
            }
        )


# =============================================================================
# GET /api/reports/stats/
# =============================================================================

class ReportStatsView(APIView):
    permission_classes = [IsAuthenticated, HasAllowedRoles]
    allowed_roles = SALES_ROLES

    def get(self, request):
        fy_id = (
            request.query_params.get("financialYearId")
            or request.query_params.get("financial_year")
        )
        m = _M()

        try:
            # Sales outstanding
            s_qs = m["SalesInvoice"].objects.exclude(status="cancelled")
            if fy_id:
                s_qs = s_qs.filter(financial_year_id=fy_id)
            sales_out = sum(
                item["balance_amount"]
                for item in (
                    _sales_invoice_financials(invoice)
                    for invoice in s_qs.prefetch_related("returns")
                )
                if item["balance_amount"] > 0.009
            )

            # Purchase outstanding
            p_qs = m["PurchaseEntry"].objects.filter(
                payment_status__in=["unpaid", "partial"]
            )
            if fy_id:
                p_qs = p_qs.filter(financial_year_id=fy_id)
            pur_out = _f(p_qs.aggregate(t=Sum("balance_amount"))["t"])

            # Counts
            total_assets = m["Asset"].objects.filter(status="active").count()
            low_stock = m["Stock"].objects.filter(
                total_quantity__lte=F("minimum_stock")
            ).count()
            total_suppliers = m["Supplier"].objects.filter(is_active=True).count()

            inv_qs = m["SalesInvoice"].objects.all()
            if fy_id:
                inv_qs = inv_qs.filter(financial_year_id=fy_id)
            total_invoices = inv_qs.count()

        except Exception as exc:
            logger.exception("Stats error: %s", exc)
            return Response({"error": str(exc)}, status=500)

        return Response(
            {
                "totalAssets": total_assets,
                "purchaseOutstanding": round(pur_out, 2),
                "salesOutstanding": round(sales_out, 2),
                "totalSuppliers": total_suppliers,
                "lowStockItems": low_stock,
                "totalInvoices": total_invoices,
            }
        )


# =============================================================================
# POST /api/reports/preview/
# =============================================================================

class PreviewReportView(APIView):
    permission_classes = [IsAuthenticated, HasAllowedRoles]
    allowed_roles = SALES_ROLES

    def post(self, request):
        rtype = request.data.get("report_type")
        df = request.data.get("date_from")
        dt = request.data.get("date_to")
        extra = request.data.get("filters") or {}
        fy_id = _fy_id(request.data)
        selected_columns = request.data.get("selected_columns") or []

        try:
            page = max(1, int(request.data.get("page", 1)))
            page_size = max(1, min(int(request.data.get("page_size", 100)), 500))
        except (TypeError, ValueError):
            page, page_size = 1, 100

        try:
            data = _build(rtype, df, dt, extra, fy_id)
        except ReportBuildError as exc:
            return Response({"error": str(exc)}, status=500)

        if data is None:
            return Response({"error": f"Unknown report type: {rtype}"}, status=400)

        data = _apply_selected_columns(data, selected_columns)

        total = len(data["rows"])
        total_pages = max(1, -(-total // page_size))
        start = (page - 1) * page_size
        end = start + page_size

        return Response(
            {
                "title": data["title"],
                "headers": data["headers"],
                "rows": data["rows"][start:end],
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
            }
        )


# =============================================================================
# POST /api/reports/generate/
# =============================================================================

class GenerateReportView(APIView):
    permission_classes = [IsAuthenticated, HasAllowedRoles]
    allowed_roles = SALES_ROLES

    def post(self, request):
        rtype = request.data.get("report_type")
        fmt = (request.data.get("format") or "pdf").lower()
        df = request.data.get("date_from")
        dt = request.data.get("date_to")
        extra = request.data.get("filters") or {}
        fy_id = _fy_id(request.data)
        selected_columns = request.data.get("selected_columns") or []

        if fmt not in ("csv", "xlsx", "pdf"):
            return Response({"error": "format must be csv, xlsx or pdf"}, status=400)

        try:
            data = _build(rtype, df, dt, extra, fy_id)
        except ReportBuildError as exc:
            return Response({"error": str(exc)}, status=500)

        if data is None:
            return Response({"error": f"Unknown report type: {rtype}"}, status=400)

        data = _apply_selected_columns(data, selected_columns)

        if fmt == "csv":
            return _render_csv(data)
        if fmt == "xlsx":
            return _render_xlsx(data)
        return _render_pdf(data, fy_id)


# =============================================================================
# Dispatch registry
# =============================================================================

_REGISTRY: dict[str, str] = {
    # Sales
    "sales-summary":        "_r_sales_summary",
    "monthly-sales":        "_r_monthly_sales",
    "daily-sales":          "_r_daily_sales",
    "sales-by-customer":    "_r_sales_by_customer",
    "sales-by-item":        "_r_sales_by_item",
    "sales-by-salesperson": "_r_sales_by_salesperson",
    "salesperson-collections": "_r_salesperson_collections",
    "sales-returns":        "_r_sales_returns",
    "sales-returns-by-customer": "_r_sales_returns_by_customer",
    "payment-collection":   "_r_payment_collection",
    "outstanding":          "_r_outstanding",
    "profit-loss":          "_r_profit_loss",
    # Purchase
    "purchase-summary":     "_r_purchase_summary",
    "purchase-by-supplier": "_r_purchase_by_supplier",
    "purchase-entries":     "_r_purchase_entries",
    "purchase-orders":      "_r_purchase_orders",
    "purchase-payments":    "_r_purchase_payments",
    "supplier-outstanding": "_r_supplier_outstanding",
    # Inventory
    "stock-summary":        "_r_stock_summary",
    "stock-low":            "_r_stock_low",
    "stock-movement":       "_r_stock_movement",
    # Directories
    "supplier-list":        "_r_supplier_list",
    "supplier-detail":      "_r_supplier_detail",
    "supplier-statement":   "_r_supplier_statement",
    "customer-list":        "_r_customer_list",
    "customer-detail":      "_r_customer_detail",
    "customer-statement":   "_r_customer_statement",
}


def _build(rtype: str, df, dt, extra: dict, fy_id) -> dict | None:
    fn_name = _REGISTRY.get(rtype or "")
    if not fn_name:
        return None
    fn = globals().get(fn_name)
    if fn is None:
        return None
    try:
        return fn(df, dt, extra, fy_id)
    except Exception as exc:
        traceback.print_exc()
        raise ReportBuildError(f"Error building '{rtype}': {exc}") from exc


def _apply_selected_columns(data: dict, selected_columns) -> dict:
    if not data or not selected_columns:
        return data

    requested = [str(column).strip() for column in selected_columns if str(column).strip()]
    if not requested:
        return data

    headers = data.get("headers") or []
    rows = data.get("rows") or []
    index_map = {header: idx for idx, header in enumerate(headers)}
    kept_headers = [header for header in headers if header in requested]
    if not kept_headers:
        return data

    kept_indices = [index_map[header] for header in kept_headers]
    filtered_rows = []
    for row in rows:
        filtered_rows.append(
            [row[idx] if idx < len(row) else "" for idx in kept_indices]
        )

    return {
        **data,
        "headers": kept_headers,
        "rows": filtered_rows,
    }


# =============================================================================
# ── SALES REPORT BUILDERS
# =============================================================================

def _r_sales_summary(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesInvoice"].objects.all()
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    ps = (extra.get("payment_status") or "").strip()
    srch = (extra.get("customer_name") or extra.get("search") or "").strip()
    if ps:
        qs = qs.filter(payment_status=ps)
    if srch:
        qs = qs.filter(customer_name__icontains=srch)

    rows = []
    for inv in qs.order_by("-invoice_date").values(
        "invoice_number",
        "customer_name",
        "invoice_date",
        "due_date",
        "subtotal",
        "disc_amount",
        "tax_amount",
        "total_amount",
        "paid_amount",
        "balance_amount",
        "payment_status",
        "status",
    ):
        rows.append(
            [
                _str(inv["invoice_number"]),
                _str(inv["customer_name"]),
                _ds(inv["invoice_date"]),
                _ds(inv["due_date"]),
                _aed(inv["subtotal"]),
                _aed(inv["disc_amount"]),
                _aed(inv["tax_amount"]),
                _aed(inv["total_amount"]),
                _aed(inv["paid_amount"]),
                _aed(inv["balance_amount"]),
                _upper(inv["payment_status"]),
                _upper(inv["status"]),
            ]
        )

    return {
        "title": "Sales Summary Report",
        "headers": [
            "Invoice #",
            "Customer",
            "Date",
            "Due Date",
            "Subtotal",
            "Discount",
            "VAT",
            "Total",
            "Paid",
            "Balance",
            "Pay. Status",
            "Status",
        ],
        "rows": rows,
    }


def _r_monthly_sales(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesInvoice"].objects.exclude(status="cancelled")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    data = list(
        qs.annotate(mo=TruncMonth("invoice_date"))
        .values("mo")
        .annotate(
            cnt=Count("id"),
            rev=Sum("total_amount"),
            tax=Sum("tax_amount"),
            disc=Sum("disc_amount"),
            paid=Sum("paid_amount"),
            bal=Sum("balance_amount"),
        )
        .order_by("mo")
    )

    rows = []
    t_cnt = t_rev = t_tax = t_disc = t_paid = t_bal = 0.0

    for r in data:
        cnt = r["cnt"] or 0
        rev = _f(r["rev"])
        tax = _f(r["tax"])
        disc = _f(r["disc"])
        paid = _f(r["paid"])
        bal = _f(r["bal"])
        t_cnt += cnt
        t_rev += rev
        t_tax += tax
        t_disc += disc
        t_paid += paid
        t_bal += bal
        rows.append(
            [
                r["mo"].strftime("%b %Y"),
                str(cnt),
                _aed(disc),
                _aed(tax),
                _aed(rev),
                _aed(paid),
                _aed(bal),
            ]
        )

    rows.append(
        [
            "TOTAL",
            str(int(t_cnt)),
            _aed(t_disc),
            _aed(t_tax),
            _aed(t_rev),
            _aed(t_paid),
            _aed(t_bal),
        ]
    )

    return {
        "title": "Monthly Sales Report",
        "headers": [
            "Month",
            "Invoices",
            "Discount",
            "VAT",
            "Total Revenue",
            "Collected",
            "Outstanding",
        ],
        "rows": rows,
    }


def _r_daily_sales(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesInvoice"].objects.exclude(status="cancelled")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    data = list(
        qs.annotate(day=TruncDate("invoice_date"))
        .values("day")
        .annotate(
            cnt=Count("id"),
            rev=Sum("total_amount"),
            tax=Sum("tax_amount"),
            paid=Sum("paid_amount"),
        )
        .order_by("day")
    )

    rows = [
        [
            _ds(r["day"]),
            str(r["cnt"]),
            _aed(r["tax"]),
            _aed(r["rev"]),
            _aed(r["paid"]),
        ]
        for r in data
    ]

    return {
        "title": "Daily Sales Report",
        "headers": ["Date", "Invoices", "VAT", "Total Revenue", "Collected"],
        "rows": rows,
    }


def _r_sales_by_customer(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesInvoice"].objects.exclude(status="cancelled")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    srch = (extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(customer_name__icontains=srch)

    data = list(
        qs.values("customer_name")
        .annotate(
            cnt=Count("id"),
            rev=Sum("total_amount"),
            paid=Sum("paid_amount"),
            bal=Sum("balance_amount"),
        )
        .order_by("-rev")
    )

    rows = [
        [
            str(i),
            _str(r["customer_name"]),
            str(r["cnt"]),
            _aed(r["rev"]),
            _aed(r["paid"]),
            _aed(r["bal"]),
        ]
        for i, r in enumerate(data, 1)
    ]

    return {
        "title": "Sales by Customer Report",
        "headers": ["#", "Customer", "Invoices", "Total", "Collected", "Outstanding"],
        "rows": rows,
    }


def _r_sales_by_item(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesInvoiceItem"].objects.select_related("invoice")
    if fy:
        qs = qs.filter(invoice__financial_year=fy)
    if f:
        qs = qs.filter(invoice__invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice__invoice_date__lte=t)

    qs = qs.exclude(invoice__status="cancelled")

    srch = (extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(item_name__icontains=srch)

    data = list(
        qs.values("item_name")
        .annotate(
            cnt=Count("invoice", distinct=True),
            qty=Sum("quantity"),
            rev=Sum("line_total"),
            tax=Sum("tax_amount"),
        )
        .order_by("-rev")
    )

    rows = [
        [
            str(i),
            _str(r["item_name"]),
            str(r["cnt"]),
            _num(r["qty"]),
            _aed(r["tax"]),
            _aed(r["rev"]),
        ]
        for i, r in enumerate(data, 1)
    ]

    return {
        "title": "Sales by Item Report",
        "headers": ["#", "Item Name", "Invoices", "Qty", "VAT", "Revenue"],
        "rows": rows,
    }


def _r_sales_by_salesperson(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesInvoice"].objects.exclude(status="cancelled")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    srch = (extra.get("salesperson_name") or extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(sales_person__name__icontains=srch)

    ps = (extra.get("payment_status") or "").strip()
    if ps:
        qs = qs.filter(payment_status=ps)

    data = list(
        qs.values("sales_person__name")
        .annotate(
            cnt=Count("id"),
            subtotal=Sum("subtotal"),
            tax=Sum("tax_amount"),
            rev=Sum("total_amount"),
            paid=Sum("paid_amount"),
            bal=Sum("balance_amount"),
        )
        .order_by("-rev", "sales_person__name")
    )

    rows = []
    for i, r in enumerate(data, 1):
        revenue = _f(r["rev"])
        invoice_count = int(r["cnt"] or 0)
        average = revenue / invoice_count if invoice_count else 0
        rows.append(
            [
                str(i),
                _str(r["sales_person__name"] or "Unassigned"),
                str(invoice_count),
                _aed(r["subtotal"]),
                _aed(r["tax"]),
                _aed(revenue),
                _aed(r["paid"]),
                _aed(r["bal"]),
                _aed(average),
            ]
        )

    return {
        "title": "Sales by Salesperson Report",
        "headers": [
            "#",
            "Salesperson",
            "Invoices",
            "Subtotal",
            "VAT",
            "Revenue",
            "Collected",
            "Outstanding",
            "Avg / Invoice",
        ],
        "rows": rows,
    }


def _r_salesperson_collections(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesPayment"].objects.select_related("sales_invoice", "sales_invoice__sales_person")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(payment_date__gte=f)
    if t:
        qs = qs.filter(payment_date__lte=t)

    srch = (extra.get("salesperson_name") or extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(sales_invoice__sales_person__name__icontains=srch)

    method = (extra.get("payment_method") or "").strip()
    if method:
        qs = qs.filter(payment_method=method)

    rows = []
    for p in qs.order_by("-payment_date").values(
        "payment_date",
        "sales_invoice__invoice_number",
        "sales_invoice__customer_name",
        "sales_invoice__sales_person__name",
        "amount",
        "payment_method",
        "reference_no",
        "notes",
    ):
        rows.append(
            [
                _ds(p["payment_date"]),
                _str(p["sales_invoice__sales_person__name"] or "Unassigned"),
                _str(p["sales_invoice__invoice_number"]),
                _str(p["sales_invoice__customer_name"]),
                _aed(p["amount"]),
                _title(p["payment_method"]),
                _str(p["reference_no"]),
                _str(p["notes"]),
            ]
        )

    return {
        "title": "Salesperson Collections Report",
        "headers": [
            "Date",
            "Salesperson",
            "Invoice #",
            "Customer",
            "Amount",
            "Method",
            "Reference",
            "Notes",
        ],
        "rows": rows,
    }


def _r_sales_returns(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesReturn"].objects.select_related(
        "customer", "sales_invoice", "warehouse", "created_by"
    ).prefetch_related("items__sales_invoice_item__invoice")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(return_date__gte=f)
    if t:
        qs = qs.filter(return_date__lte=t)

    status_filter = (extra.get("status") or "").strip()
    if status_filter:
        qs = qs.filter(status=status_filter)

    srch = (extra.get("search") or extra.get("customer_name") or "").strip()
    if srch:
        qs = qs.filter(
            Q(return_number__icontains=srch)
            | Q(customer__display_name__icontains=srch)
            | Q(sales_invoice__invoice_number__icontains=srch)
            | Q(items__sales_invoice_item__invoice__invoice_number__icontains=srch)
        ).distinct()

    rows = []
    for ret in qs.order_by("-return_date", "-created_at"):
        invoice_numbers = sorted(
            {
                item.sales_invoice_item.invoice.invoice_number
                for item in ret.items.all()
                if item.sales_invoice_item_id and item.sales_invoice_item.invoice_id
            }
        )
        dispositions = sorted(
            {
                item.get_disposition_display()
                for item in ret.items.all()
                if getattr(item, "disposition", "")
            }
        )
        rows.append(
            [
                _str(ret.return_number),
                _ds(ret.return_date),
                _customer_name(ret.customer),
                ", ".join(invoice_numbers) if invoice_numbers else _str(ret.sales_invoice.invoice_number if ret.sales_invoice_id else ""),
                _str(ret.warehouse.name if ret.warehouse_id else ""),
                str(ret.items.count()),
                ", ".join(dispositions) if dispositions else "—",
                _aed(ret.subtotal),
                _aed(ret.tax_amount),
                _aed(ret.total_amount),
                _upper(ret.status),
                "YES" if ret.stock_posted else "NO",
            ]
        )

    return {
        "title": "Sales Returns Report",
        "headers": [
            "Return #",
            "Return Date",
            "Customer",
            "Invoice #",
            "Warehouse",
            "Lines",
            "Actions",
            "Subtotal",
            "VAT",
            "Credit Total",
            "Status",
            "Stock Posted",
        ],
        "rows": rows,
    }


def _r_sales_returns_by_customer(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesReturn"].objects.select_related("customer")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(return_date__gte=f)
    if t:
        qs = qs.filter(return_date__lte=t)

    status_filter = (extra.get("status") or "").strip()
    if status_filter:
        qs = qs.filter(status=status_filter)

    srch = (extra.get("search") or extra.get("customer_name") or "").strip()
    if srch:
        qs = qs.filter(
            Q(customer__display_name__icontains=srch)
            | Q(customer__company_name__icontains=srch)
            | Q(customer__first_name__icontains=srch)
            | Q(customer__last_name__icontains=srch)
        )

    data = list(
        qs.values("customer_id")
        .annotate(
            cnt=Count("id"),
            subtotal=Sum("subtotal"),
            tax=Sum("tax_amount"),
            total=Sum("total_amount"),
            confirmed=Count("id", filter=Q(status="confirmed")),
            draft=Count("id", filter=Q(status="draft")),
            cancelled=Count("id", filter=Q(status="cancelled")),
        )
        .order_by("-total")
    )

    customer_ids = [row["customer_id"] for row in data if row["customer_id"]]
    customer_map = m["Customer"].objects.in_bulk(customer_ids) if customer_ids else {}

    rows = []
    for i, row in enumerate(data, 1):
        customer = customer_map.get(row["customer_id"])
        rows.append(
            [
                str(i),
                _customer_name(customer),
                str(row["cnt"] or 0),
                str(row["confirmed"] or 0),
                str(row["draft"] or 0),
                str(row["cancelled"] or 0),
                _aed(row["subtotal"]),
                _aed(row["tax"]),
                _aed(row["total"]),
            ]
        )

    return {
        "title": "Sales Returns by Customer Report",
        "headers": [
            "#",
            "Customer",
            "Returns",
            "Confirmed",
            "Draft",
            "Cancelled",
            "Subtotal",
            "VAT",
            "Credit Total",
        ],
        "rows": rows,
    }


def _r_payment_collection(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesPayment"].objects.select_related("sales_invoice")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(payment_date__gte=f)
    if t:
        qs = qs.filter(payment_date__lte=t)

    srch = (extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(
            Q(sales_invoice__customer_name__icontains=srch)
            | Q(sales_invoice__invoice_number__icontains=srch)
            | Q(reference_no__icontains=srch)
        )

    rows = []
    for p in qs.order_by("-payment_date").values(
        "payment_date",
        "sales_invoice__invoice_number",
        "sales_invoice__customer_name",
        "amount",
        "payment_method",
        "reference_no",
        "notes",
    ):
        rows.append(
            [
                _ds(p["payment_date"]),
                _str(p["sales_invoice__invoice_number"]),
                _str(p["sales_invoice__customer_name"]),
                _aed(p["amount"]),
                _title(p["payment_method"]),
                _str(p["reference_no"]),
                _str(p["notes"]),
            ]
        )

    return {
        "title": "Payment Collection Report",
        "headers": [
            "Date",
            "Invoice #",
            "Customer",
            "Amount",
            "Method",
            "Reference",
            "Notes",
        ],
        "rows": rows,
    }


def _r_outstanding(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["SalesInvoice"].objects.filter(
        payment_status__in=["unpaid", "partial"]
    ).exclude(status="cancelled")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    srch = (extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(customer_name__icontains=srch)

    today = _date.today()
    rows = []

    for inv in qs.order_by("due_date").values(
        "invoice_number",
        "customer_name",
        "invoice_date",
        "due_date",
        "total_amount",
        "paid_amount",
        "balance_amount",
        "payment_status",
    ):
        due = inv["due_date"]
        overdue_days = (today - due).days if due else 0
        overdue_label = (
            f"{overdue_days}d overdue" if overdue_days > 0 else "On time"
        )
        rows.append(
            [
                _str(inv["invoice_number"]),
                _str(inv["customer_name"]),
                _ds(inv["invoice_date"]),
                _ds(due),
                _aed(inv["total_amount"]),
                _aed(inv["paid_amount"]),
                _aed(inv["balance_amount"]),
                _upper(inv["payment_status"]),
                overdue_label,
            ]
        )

    return {
        "title": "Outstanding Dues Report",
        "headers": [
            "Invoice #",
            "Customer",
            "Inv. Date",
            "Due Date",
            "Total",
            "Paid",
            "Balance",
            "Status",
            "Overdue",
        ],
        "rows": rows,
    }


def _r_profit_loss(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    s_qs = m["SalesInvoice"].objects.exclude(status="cancelled")
    p_qs = m["PurchaseEntry"].objects.all()

    if fy:
        s_qs = s_qs.filter(financial_year=fy)
        p_qs = p_qs.filter(financial_year=fy)
    if f:
        s_qs = s_qs.filter(invoice_date__gte=f)
        p_qs = p_qs.filter(invoice_date__gte=f)
    if t:
        s_qs = s_qs.filter(invoice_date__lte=t)
        p_qs = p_qs.filter(invoice_date__lte=t)

    sales_by_mo = {
        r["mo"].strftime("%b %Y"): _f(r["rev"])
        for r in s_qs.annotate(mo=TruncMonth("invoice_date"))
        .values("mo")
        .annotate(rev=Sum("total_amount"))
        .order_by("mo")
    }
    cost_by_mo = {
        r["mo"].strftime("%b %Y"): _f(r["cost"])
        for r in p_qs.annotate(mo=TruncMonth("invoice_date"))
        .values("mo")
        .annotate(cost=Sum("total_amount"))
        .order_by("mo")
    }

    months = sorted(set(sales_by_mo) | set(cost_by_mo))
    rows = []
    t_s = t_p = 0.0

    for mo in months:
        s = sales_by_mo.get(mo, 0.0)
        p = cost_by_mo.get(mo, 0.0)
        gp = s - p
        t_s += s
        t_p += p
        rows.append(
            [
                mo,
                _aed(s),
                _aed(p),
                _aed(gp),
                _pct((gp / s * 100) if s else 0),
            ]
        )

    t_gp = t_s - t_p
    rows.append(
        [
            "TOTAL",
            _aed(t_s),
            _aed(t_p),
            _aed(t_gp),
            _pct((t_gp / t_s * 100) if t_s else 0),
        ]
    )

    return {
        "title": "Profit & Loss Report",
        "headers": [
            "Month",
            "Sales Revenue",
            "Purchase Cost",
            "Gross Profit",
            "Margin %",
        ],
        "rows": rows,
    }


# =============================================================================
# ── PURCHASE REPORT BUILDERS
# =============================================================================

def _r_purchase_summary(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["PurchaseEntry"].objects.select_related("vendor")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    ps = (extra.get("payment_status") or "").strip()
    srch = (extra.get("search") or "").strip()
    if ps:
        qs = qs.filter(payment_status=ps)
    if srch:
        qs = qs.filter(_supplier_fk_q(srch))

    rows = []
    for e in qs.order_by("-invoice_date"):
        rows.append(
            [
                _str(e.entry_number),
                _str(e.vendor_invoice_no),
                _supplier_name(e.vendor),
                _ds(e.invoice_date),
                _ds(e.due_date),
                _aed(e.total_amount),
                _aed(e.paid_amount),
                _aed(e.balance_amount),
                _upper(e.payment_status),
                "Yes" if e.is_received else "No",
            ]
        )

    return {
        "title": "Purchase Entry Summary",
        "headers": [
            "Entry #",
            "Supplier Inv.",
            "Supplier",
            "Date",
            "Due",
            "Total",
            "Paid",
            "Balance",
            "Pay. Status",
            "Received",
        ],
        "rows": rows,
    }


def _r_purchase_by_supplier(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["PurchaseEntry"].objects.all()
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    srch = (extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(_supplier_fk_q(srch))

    data = list(
        qs.values("vendor_id")
        .annotate(
            cnt=Count("id"),
            tot=Sum("total_amount"),
            paid=Sum("paid_amount"),
            bal=Sum("balance_amount"),
        )
        .order_by("-tot")
    )

    sup_ids = [r["vendor_id"] for r in data if r["vendor_id"]]
    supplier_map = m["Supplier"].objects.in_bulk(sup_ids) if sup_ids else {}

    rows = [
        [
            str(i),
            _supplier_name(supplier_map.get(r["vendor_id"])),
            str(r["cnt"]),
            _aed(r["tot"]),
            _aed(r["paid"]),
            _aed(r["bal"]),
        ]
        for i, r in enumerate(data, 1)
    ]

    return {
        "title": "Purchase by Supplier Report",
        "headers": ["#", "Supplier", "Entries", "Total", "Paid", "Outstanding"],
        "rows": rows,
    }


def _r_purchase_entries(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["PurchaseEntry"].objects.select_related("vendor")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(invoice_date__gte=f)
    if t:
        qs = qs.filter(invoice_date__lte=t)

    srch = (extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(_supplier_fk_q(srch))

    rows = []
    for e in qs.order_by("-invoice_date"):
        rows.append(
            [
                _str(e.entry_number),
                _supplier_name(e.vendor),
                _ds(e.invoice_date),
                _str(e.vendor_invoice_no),
                _aed(e.total_amount),
                _aed(e.paid_amount),
                _aed(e.balance_amount),
                _upper(e.payment_status),
                "Yes" if e.is_received else "Pending",
                _ds(e.received_at) if e.is_received else "—",
            ]
        )

    return {
        "title": "Purchase Entries (GRN) Report",
        "headers": [
            "Entry #",
            "Supplier",
            "Date",
            "Supplier Inv.",
            "Total",
            "Paid",
            "Balance",
            "Status",
            "Received",
            "Received On",
        ],
        "rows": rows,
    }


def _r_purchase_orders(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["PurchaseOrder"].objects.select_related("vendor")
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(order_date__gte=f)
    if t:
        qs = qs.filter(order_date__lte=t)

    ps = (extra.get("status") or "").strip()
    if ps:
        qs = qs.filter(status=ps)

    rows = [
        [
            _str(o.po_number),
            _supplier_name(o.vendor),
            _ds(o.order_date),
            _aed(o.total_amount),
            _upper(o.status),
        ]
        for o in qs.order_by("-order_date")
    ]

    return {
        "title": "Purchase Orders Report",
        "headers": ["PO #", "Supplier", "Order Date", "Amount", "Status"],
        "rows": rows,
    }


def _r_purchase_payments(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    qs = m["PurchasePayment"].objects.select_related(
        "purchase_entry", "purchase_entry__vendor"
    )
    if fy:
        qs = qs.filter(financial_year=fy)
    if f:
        qs = qs.filter(payment_date__gte=f)
    if t:
        qs = qs.filter(payment_date__lte=t)

    srch = (extra.get("search") or "").strip()
    if srch:
        qs = qs.filter(
            Q(purchase_entry__vendor__display_name__icontains=srch)
            | Q(purchase_entry__vendor__company_name__icontains=srch)
            | Q(purchase_entry__entry_number__icontains=srch)
            | Q(reference_no__icontains=srch)
        )

    rows = []
    for p in qs.order_by("-payment_date"):
        entry = p.purchase_entry
        rows.append(
            [
                _ds(p.payment_date),
                _supplier_name(entry.vendor if entry else None),
                _str(entry.entry_number) if entry else "—",
                _aed(p.amount),
                _title(p.payment_method),
                _str(p.reference_no),
            ]
        )

    return {
        "title": "Purchase Payments Report",
        "headers": ["Date", "Supplier", "Entry #", "Amount", "Method", "Reference"],
        "rows": rows,
    }


def _r_supplier_outstanding(df, dt, extra, fy_id):
    m = _M()
    f, t, fy = _bounds(fy_id, df, dt)

    srch = (extra.get("search") or "").strip()
    suppliers = m["Supplier"].objects.filter(is_active=True).order_by(
        "display_name", "company_name", "first_name"
    )
    if srch:
        suppliers = suppliers.filter(_supplier_obj_q(srch))

    rows = []
    for sup in suppliers:
        qs = m["PurchaseEntry"].objects.filter(
            vendor=sup, payment_status__in=["unpaid", "partial"]
        )
        if fy:
            qs = qs.filter(financial_year=fy)
        if f:
            qs = qs.filter(invoice_date__gte=f)
        if t:
            qs = qs.filter(invoice_date__lte=t)

        agg = qs.aggregate(
            tot=Sum("total_amount"),
            paid=Sum("paid_amount"),
            bal=Sum("balance_amount"),
            cnt=Count("id"),
        )
        if not agg["cnt"]:
            continue

        phone = sup.phone or getattr(sup, "mobile", "") or "—"
        trn = sup.trn or getattr(sup, "gstin", "") or "—"
        rows.append(
            [
                _supplier_name(sup),
                phone,
                trn,
                str(agg["cnt"]),
                _aed(agg["tot"]),
                _aed(agg["paid"]),
                _aed(agg["bal"]),
            ]
        )

    return {
        "title": "Supplier Outstanding Report",
        "headers": [
            "Supplier",
            "Phone",
            "TRN",
            "Open Entries",
            "Total Invoiced",
            "Paid",
            "Balance Due",
        ],
        "rows": rows,
    }


# =============================================================================
# ── INVENTORY REPORT BUILDERS
# =============================================================================

def _r_stock_summary(df, dt, extra, fy_id):
    m = _M()

    qs = m["Stock"].objects.select_related("item", "warehouse")

    wh = (extra.get("warehouse") or "").strip()
    srch = (extra.get("search") or "").strip()
    if wh:
        qs = qs.filter(warehouse__name__icontains=wh)
    if srch:
        qs = qs.filter(
            Q(item__name__icontains=srch) | Q(item__sku__icontains=srch)
        )

    rows = []
    for s in qs.order_by("warehouse__name", "item__name"):
        avail = s.available_quantity
        if avail <= 0:
            status = "OUT OF STOCK"
        elif s.is_low_stock:
            status = "LOW STOCK"
        else:
            status = "OK"

        val_pur = avail * _f(s.item.purchase_cost)
        val_sel = avail * _f(s.item.selling_price)
        cat_name = _title(s.item.item_type)

        rows.append(
            [
                _str(s.item.asset_code),
                _str(s.item.name),
                cat_name,
                _str(s.warehouse.name),
                str(s.total_quantity),
                str(s.damaged_quantity),
                str(avail),
                str(s.minimum_stock),
                status,
                _aed(val_pur),
                _aed(val_sel),
            ]
        )

    return {
        "title": "Stock Summary Report",
        "headers": [
            "Asset Code",
            "Asset Name",
            "Category",
            "Warehouse",
            "Total",
            "Damaged",
            "Available",
            "Min Level",
            "Status",
            "Value (Purchase)",
            "Value (Selling)",
        ],
        "rows": rows,
    }


def _r_stock_low(df, dt, extra, fy_id):
    m = _M()

    qs = m["Stock"].objects.select_related("item", "warehouse").filter(
        total_quantity__lte=F("minimum_stock")
    )

    wh = (extra.get("warehouse") or "").strip()
    srch = (extra.get("search") or "").strip()
    if wh:
        qs = qs.filter(warehouse__name__icontains=wh)
    if srch:
        qs = qs.filter(
            Q(item__name__icontains=srch) | Q(item__sku__icontains=srch)
        )

    rows = []
    for s in qs.order_by("total_quantity"):
        avail = s.available_quantity
        cat_name = _title(s.item.item_type)
        shortfall = max(0, s.minimum_stock - avail)
        rows.append(
            [
                _str(s.item.asset_code),
                _str(s.item.name),
                cat_name,
                _str(s.warehouse.name),
                str(avail),
                str(s.minimum_stock),
                str(shortfall),
                "OUT OF STOCK" if avail <= 0 else "LOW STOCK",
            ]
        )

    return {
        "title": "Low / Out-of-Stock Report",
        "headers": [
            "Asset Code",
            "Asset Name",
            "Category",
            "Warehouse",
            "Available",
            "Min Level",
            "Shortfall",
            "Alert",
        ],
        "rows": rows,
    }


def _r_stock_movement(df, dt, extra, fy_id):
    m = _M()
    f, t, _fy = _bounds(fy_id, df, dt)

    qs = m["StockHistory"].objects.select_related("item", "warehouse", "performed_by")
    if f:
        qs = qs.filter(created_at__date__gte=f)
    if t:
        qs = qs.filter(created_at__date__lte=t)

    wh = (extra.get("warehouse") or "").strip()
    srch = (extra.get("search") or "").strip()
    mv = (extra.get("movement_type") or "").strip()

    if wh:
        qs = qs.filter(warehouse__name__icontains=wh)
    if srch:
        qs = qs.filter(
            Q(item__name__icontains=srch) | Q(item__sku__icontains=srch)
        )
    if mv:
        qs = qs.filter(movement_type=mv)

    rows = []
    for h in qs.order_by("-created_at")[:2000]:
        performed = "System"
        if h.performed_by:
            performed = getattr(h.performed_by, "name", str(h.performed_by))
        rows.append(
            [
                _dts(h.created_at),
                _str(h.item.asset_code),
                _str(h.item.name),
                _str(h.warehouse.name),
                h.get_movement_type_display(),
                str(h.quantity),
                str(h.balance_after),
                _str(h.reason),
                performed,
            ]
        )

    return {
        "title": "Stock Movement Log",
        "headers": [
            "Date / Time",
            "Asset Code",
            "Asset Name",
            "Warehouse",
            "Movement Type",
            "Qty",
            "Balance After",
            "Reason",
            "Performed By",
        ],
        "rows": rows,
    }


# =============================================================================
# ── DIRECTORY BUILDERS
# =============================================================================

def _r_supplier_list(df, dt, extra, fy_id):
    m = _M()

    srch = (extra.get("search") or "").strip()
    qs = m["Supplier"].objects.filter(is_active=True).order_by(
        "display_name", "company_name", "first_name"
    )
    if srch:
        qs = qs.filter(_supplier_obj_q(srch))

    rows = []
    for i, s in enumerate(qs, 1):
        phone = s.phone or getattr(s, "mobile", "") or "—"
        trn = s.trn or getattr(s, "gstin", "") or "—"
        pan = getattr(s, "pan", "") or "—"
        rows.append(
            [
                str(i),
                _supplier_name(s),
                phone,
                _str(s.email) if s.email else "—",
                trn,
                pan,
                _aed(s.outstanding),
                _aed(s.credit_limit),
            ]
        )

    return {
        "title": "Supplier Directory",
        "headers": ["#", "Name", "Phone", "Email", "TRN", "PAN", "Outstanding", "Credit Limit"],
        "rows": rows,
    }


def _r_supplier_detail(df, dt, extra, fy_id):
    m = _M()
    supplier_id = extra.get("supplier_id") or extra.get("supplierId")
    search = (extra.get("search") or extra.get("supplier_name") or "").strip()

    f, t, fy = _bounds(fy_id, df, dt)
    supplier_qs = m["Supplier"].objects.all().order_by(
        "display_name", "company_name", "first_name", "last_name"
    )
    if supplier_id:
        supplier_qs = supplier_qs.filter(id=supplier_id)
    elif search:
        supplier_qs = supplier_qs.filter(_supplier_obj_q(search))

    suppliers = list(supplier_qs)
    if supplier_id and not suppliers:
        raise ReportBuildError("Supplier not found.")
    if search and not suppliers:
        raise ReportBuildError("No suppliers found for the selected search.")

    rows = []
    for supplier in suppliers:
        orders_qs = m["PurchaseOrder"].objects.filter(vendor=supplier).select_related(
            "financial_year"
        )
        entries_qs = m["PurchaseEntry"].objects.filter(vendor=supplier).select_related(
            "financial_year", "purchase_order"
        )
        payments_qs = m["PurchasePayment"].objects.filter(
            purchase_entry__vendor=supplier
        ).select_related("purchase_entry", "financial_year")

        if fy:
            orders_qs = orders_qs.filter(financial_year=fy)
            entries_qs = entries_qs.filter(financial_year=fy)
            payments_qs = payments_qs.filter(financial_year=fy)
        if f:
            orders_qs = orders_qs.filter(order_date__gte=f)
            entries_qs = entries_qs.filter(invoice_date__gte=f)
            payments_qs = payments_qs.filter(payment_date__gte=f)
        if t:
            orders_qs = orders_qs.filter(order_date__lte=t)
            entries_qs = entries_qs.filter(invoice_date__lte=t)
            payments_qs = payments_qs.filter(payment_date__lte=t)

        base_supplier = [
            "PROFILE",
            _supplier_name(supplier),
            _str(getattr(supplier, "company_name", "")),
            _str(getattr(supplier, "email", "")),
            _str(getattr(supplier, "phone", "")),
            _str(getattr(supplier, "mobile", "")),
            _str(getattr(supplier, "trn", "") or getattr(supplier, "gstin", "")),
            _str(getattr(supplier, "pan", "")),
            _str(getattr(supplier, "currency", "")),
            _title(getattr(supplier, "payment_terms", "")),
            _aed(getattr(supplier, "credit_limit", 0)),
            _aed(getattr(supplier, "outstanding", 0)),
        ]
        rows.append(base_supplier + ["—"] * 17)

        for order in orders_qs.order_by("-order_date", "-created_at"):
            rows.append(
                base_supplier
                + [
                    _str(order.po_number),
                    _ds(order.order_date),
                    _ds(order.expected_date),
                    _aed(order.total_amount),
                    _upper(order.status),
                    _str(order.reference_no),
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    _str(order.notes),
                ]
            )

        for entry in entries_qs.order_by("-invoice_date", "-created_at"):
            rows.append(
                base_supplier
                + [
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    _str(entry.entry_number),
                    _ds(entry.invoice_date),
                    _ds(entry.due_date),
                    _aed(entry.total_amount),
                    _aed(entry.paid_amount),
                    _aed(entry.balance_amount),
                    _upper(entry.payment_status),
                    "—",
                    "—",
                    _str(entry.notes),
                ]
            )

        for payment in payments_qs.order_by("-payment_date", "-created_at"):
            rows.append(
                base_supplier
                + [
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    _str(payment.purchase_entry.entry_number),
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    _ds(payment.payment_date),
                    _aed(payment.amount),
                    _title(payment.payment_method),
                    _str(payment.reference_no),
                    _str(payment.notes),
                ]
            )

    return {
        "title": (
            f"Supplier Detail - {_supplier_name(suppliers[0])}"
            if supplier_id and suppliers
            else "Supplier Detail"
        ),
        "headers": [
            "Section",
            "Supplier",
            "Company",
            "Email",
            "Phone",
            "Mobile",
            "TRN",
            "PAN",
            "Currency",
            "Payment Terms",
            "Credit Limit",
            "Outstanding",
            "PO #",
            "PO Date",
            "Expected Date",
            "PO Total",
            "PO Status",
            "PO Reference",
            "Entry #",
            "Invoice Date",
            "Due Date",
            "Entry Total",
            "Paid Amount",
            "Balance Amount",
            "Payment Status",
            "Payment Date",
            "Payment Amount",
            "Payment Method",
            "Payment Reference",
            "Notes",
        ],
        "rows": rows,
    }


def _r_customer_list(df, dt, extra, fy_id):
    m = _M()

    srch = (extra.get("search") or "").strip()
    qs = m["Customer"].objects.filter(is_active=True).order_by(
        "display_name", "company_name", "first_name"
    )
    if srch:
        qs = qs.filter(_customer_obj_q(srch))

    rows = []
    for i, c in enumerate(qs, 1):
        phone = c.phone or getattr(c, "mobile", "") or "—"
        ctype = "—"
        if hasattr(c, "get_customer_type_display"):
            ctype = c.get_customer_type_display()
        elif hasattr(c, "customer_type"):
            ctype = _title(c.customer_type)
        rows.append(
            [
                str(i),
                _customer_name(c),
                ctype,
                phone,
                _str(c.email) if c.email else "—",
                _str(c.trn) if c.trn else "—",
                _aed(c.outstanding),
                _aed(c.credit_limit),
            ]
        )

    return {
        "title": "Customer Directory",
        "headers": ["#", "Name", "Type", "Phone", "Email", "TRN", "Outstanding", "Credit Limit"],
        "rows": rows,
    }


def _r_customer_detail(df, dt, extra, fy_id):
    m = _M()
    customer_id = extra.get("customer_id") or extra.get("customerId")
    search = (extra.get("search") or extra.get("customer_name") or "").strip()

    f, t, fy = _bounds(fy_id, df, dt)
    customer_qs = m["Customer"].objects.all().order_by(
        "display_name", "company_name", "first_name", "last_name"
    )
    if customer_id:
        customer_qs = customer_qs.filter(id=customer_id)
    elif search:
        customer_qs = customer_qs.filter(_customer_obj_q(search))

    customers = list(customer_qs)
    if customer_id and not customers:
        raise ReportBuildError("Customer not found.")
    if search and not customers:
        raise ReportBuildError("No customers found for the selected search.")

    rows = []
    for customer in customers:
        invoices_qs = m["SalesInvoice"].objects.filter(customer=customer).select_related(
            "financial_year", "sales_person"
        ).prefetch_related("returns", "payments")
        payments_qs = m["SalesPayment"].objects.filter(
            sales_invoice__customer=customer
        ).select_related("sales_invoice", "financial_year")
        returns_qs = m["SalesReturn"].objects.filter(customer=customer).select_related(
            "sales_invoice", "warehouse"
        )

        if fy:
            invoices_qs = invoices_qs.filter(financial_year=fy)
            payments_qs = payments_qs.filter(financial_year=fy)
            returns_qs = returns_qs.filter(financial_year=fy)
        if f:
            invoices_qs = invoices_qs.filter(invoice_date__gte=f)
            payments_qs = payments_qs.filter(payment_date__gte=f)
            returns_qs = returns_qs.filter(return_date__gte=f)
        if t:
            invoices_qs = invoices_qs.filter(invoice_date__lte=t)
            payments_qs = payments_qs.filter(payment_date__lte=t)
            returns_qs = returns_qs.filter(return_date__lte=t)

        base_customer = [
            "PROFILE",
            _customer_name(customer),
            _str(customer.company_name),
            _title(customer.customer_type),
            _str(customer.email),
            _str(customer.phone),
            _str(customer.mobile),
            _str(customer.trn),
            _str(customer.place_of_supply),
            _str(customer.currency),
            _title(customer.payment_terms),
            _aed(customer.credit_limit),
            _aed(customer.outstanding),
        ]
        rows.append(base_customer + ["—"] * 17)

        for inv in invoices_qs.order_by("-invoice_date", "-created_at"):
            financials = _sales_invoice_financials(inv)
            rows.append(
                base_customer
                + [
                    _str(inv.invoice_number),
                    _ds(inv.invoice_date),
                    _ds(inv.due_date),
                    _aed(financials["net_total"]),
                    _aed(financials["paid_amount"]),
                    _aed(financials["balance_amount"]),
                    _upper(financials["payment_status"]),
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    _str(inv.notes),
                ]
            )

        for payment in payments_qs.order_by("-payment_date", "-created_at"):
            rows.append(
                base_customer
                + [
                    _str(payment.sales_invoice.invoice_number),
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    _ds(payment.payment_date),
                    _aed(payment.amount),
                    _title(payment.payment_method),
                    _str(payment.reference_no),
                    "—",
                    "—",
                    "—",
                    "—",
                    _str(payment.notes),
                ]
            )

        for ret in returns_qs.order_by("-return_date", "-created_at"):
            rows.append(
                base_customer
                + [
                    _str(ret.sales_invoice.invoice_number if ret.sales_invoice_id else ""),
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    _str(ret.return_number),
                    _ds(ret.return_date),
                    _aed(ret.total_amount),
                    _upper(ret.status),
                    _str(ret.warehouse.name if ret.warehouse_id else ""),
                    _str(ret.notes or ret.reason),
                ]
            )

    return {
        "title": (
            f"Customer Detail - {_customer_name(customers[0])}"
            if customer_id and customers
            else "Customer Detail"
        ),
        "headers": [
            "Section",
            "Customer",
            "Company",
            "Customer Type",
            "Email",
            "Phone",
            "Mobile",
            "TRN",
            "Place Of Supply",
            "Currency",
            "Payment Terms",
            "Credit Limit",
            "Outstanding",
            "Invoice #",
            "Invoice Date",
            "Due Date",
            "Invoice Total",
            "Paid Amount",
            "Balance Amount",
            "Payment Status",
            "Payment Date",
            "Payment Amount",
            "Payment Method",
            "Payment Reference",
            "Return #",
            "Return Date",
            "Return Amount",
            "Return Status",
            "Warehouse",
            "Notes",
        ],
        "rows": rows,
    }


def _r_customer_statement(df, dt, extra, fy_id):
    m = _M()
    customer_id = extra.get("customer_id") or extra.get("customerId")
    search = (extra.get("search") or extra.get("customer_name") or "").strip()

    f, t, fy = _bounds(fy_id, df, dt)
    customer_qs = m["Customer"].objects.all().order_by(
        "display_name", "company_name", "first_name", "last_name"
    )
    if customer_id:
        customer_qs = customer_qs.filter(id=customer_id)
    elif search:
        customer_qs = customer_qs.filter(_customer_obj_q(search))

    customer = customer_qs.first()
    if customer_id and not customer:
        raise ReportBuildError("Customer not found.")
    if search and not customer:
        raise ReportBuildError("No customers found for the selected search.")
    if not customer:
        raise ReportBuildError("Customer is required for statement download.")

    invoices_qs = m["SalesInvoice"].objects.filter(customer=customer).prefetch_related(
        "returns"
    )
    payments_qs = m["SalesPayment"].objects.filter(
        sales_invoice__customer=customer
    ).select_related("sales_invoice")
    returns_qs = m["SalesReturn"].objects.filter(
        customer=customer, status="confirmed"
    )

    if fy:
        invoices_qs = invoices_qs.filter(financial_year=fy)
        payments_qs = payments_qs.filter(financial_year=fy)
        returns_qs = returns_qs.filter(financial_year=fy)
    if f:
        invoices_qs = invoices_qs.filter(invoice_date__gte=f)
        payments_qs = payments_qs.filter(payment_date__gte=f)
        returns_qs = returns_qs.filter(return_date__gte=f)
    if t:
        invoices_qs = invoices_qs.filter(invoice_date__lte=t)
        payments_qs = payments_qs.filter(payment_date__lte=t)
        returns_qs = returns_qs.filter(return_date__lte=t)

    rows = []
    for inv in invoices_qs.order_by("invoice_date", "created_at"):
        financials = _sales_invoice_financials(inv)
        rows.append(
            {
                "date": inv.invoice_date,
                "type": "Invoice",
                "reference": _str(inv.invoice_number),
                "note": _str(inv.status),
                "debit": Decimal(str(financials["net_total"])),
                "credit": Decimal("0"),
            }
        )
    for payment in payments_qs.order_by("payment_date", "created_at"):
        rows.append(
            {
                "date": payment.payment_date,
                "type": "Payment",
                "reference": _str(payment.sales_invoice.invoice_number if payment.sales_invoice_id else ""),
                "note": _str(payment.reference_no or payment.payment_method),
                "debit": Decimal("0"),
                "credit": Decimal(str(payment.amount or 0)),
            }
        )
    for ret in returns_qs.order_by("return_date", "created_at"):
        rows.append(
            {
                "date": ret.return_date,
                "type": "Sales Return",
                "reference": _str(ret.return_number),
                "note": _str(ret.invoice_number or ret.reason or ret.notes),
                "debit": Decimal("0"),
                "credit": Decimal(str(ret.total_amount or 0)),
            }
        )

    rows.sort(key=lambda row: (row["date"], row["type"], row["reference"]))

    running = Decimal("0")
    rendered_rows = []
    for row in rows:
        running += row["debit"] - row["credit"]
        rendered_rows.append(
            [
                _ds(row["date"]),
                row["type"],
                row["reference"],
                row["note"],
                _aed(row["debit"]) if row["debit"] else "—",
                _aed(row["credit"]) if row["credit"] else "—",
                _aed(running),
            ]
        )

    title = f"Customer Statement - {_customer_name(customer)}"
    if f or t:
        title += f" ({_str(_ds(datetime.fromisoformat(str(f)).date()) if f else 'Start')} to {_str(_ds(datetime.fromisoformat(str(t)).date()) if t else 'End')})"

    return {
        "title": title,
        "headers": [
            "Date",
            "Type",
            "Reference",
            "Note",
            "Debit",
            "Credit",
            "Balance",
        ],
        "rows": rendered_rows,
    }


def _r_supplier_statement(df, dt, extra, fy_id):
    m = _M()
    supplier_id = extra.get("supplier_id") or extra.get("supplierId") or extra.get("vendor_id")
    search = (extra.get("search") or extra.get("supplier_name") or extra.get("vendor_name") or "").strip()

    f, t, fy = _bounds(fy_id, df, dt)
    supplier_qs = m["Supplier"].objects.all().order_by(
        "display_name", "company_name", "first_name", "last_name"
    )
    if supplier_id:
        supplier_qs = supplier_qs.filter(id=supplier_id)
    elif search:
        supplier_qs = supplier_qs.filter(_supplier_obj_q(search))

    supplier = supplier_qs.first()
    if supplier_id and not supplier:
        raise ReportBuildError("Supplier not found.")
    if search and not supplier:
        raise ReportBuildError("No suppliers found for the selected search.")
    if not supplier:
        raise ReportBuildError("Supplier is required for statement download.")

    entries_qs = m["PurchaseEntry"].objects.filter(vendor=supplier)
    payments_qs = m["PurchasePayment"].objects.filter(purchase_entry__vendor=supplier).select_related("purchase_entry")
    petty_cash_qs = m["PettyCashEntry"].objects.filter(
        related_party_type="vendor",
        vendor=supplier,
    )

    if fy:
        entries_qs = entries_qs.filter(financial_year=fy)
        payments_qs = payments_qs.filter(financial_year=fy)
        petty_cash_qs = petty_cash_qs.filter(financial_year=fy)
    if f:
        entries_qs = entries_qs.filter(invoice_date__gte=f)
        payments_qs = payments_qs.filter(payment_date__gte=f)
        petty_cash_qs = petty_cash_qs.filter(transaction_date__gte=f)
    if t:
        entries_qs = entries_qs.filter(invoice_date__lte=t)
        payments_qs = payments_qs.filter(payment_date__lte=t)
        petty_cash_qs = petty_cash_qs.filter(transaction_date__lte=t)

    rows = []
    for entry in entries_qs.order_by("invoice_date", "created_at"):
        rows.append(
            {
                "date": entry.invoice_date,
                "type": "Purchase Entry",
                "reference": _str(entry.entry_number),
                "note": _str(entry.vendor_invoice_no or entry.payment_status),
                "debit": Decimal(str(entry.total_amount or 0)),
                "credit": Decimal("0"),
            }
        )
    for payment in payments_qs.order_by("payment_date", "created_at"):
        rows.append(
            {
                "date": payment.payment_date,
                "type": "Payment",
                "reference": _str(payment.purchase_entry.entry_number if payment.purchase_entry_id else ""),
                "note": _str(payment.reference_no or payment.payment_method),
                "debit": Decimal("0"),
                "credit": Decimal(str(payment.amount or 0)),
            }
        )
    for entry in petty_cash_qs.order_by("transaction_date", "created_at"):
        amount = Decimal(str(entry.amount or 0))
        rows.append(
            {
                "date": entry.transaction_date,
                "type": "Petty Cash",
                "reference": _str(entry.description),
                "note": _str(getattr(entry, "get_category_display", lambda: entry.category)()),
                "debit": amount if entry.entry_type == "debit" else Decimal("0"),
                "credit": amount if entry.entry_type == "credit" else Decimal("0"),
            }
        )

    rows.sort(key=lambda row: (row["date"], row["type"], row["reference"]))

    running = Decimal("0")
    rendered_rows = []
    for row in rows:
        running += row["debit"] - row["credit"]
        rendered_rows.append(
            [
                _ds(row["date"]),
                row["type"],
                row["reference"],
                row["note"],
                _aed(row["debit"]) if row["debit"] else "—",
                _aed(row["credit"]) if row["credit"] else "—",
                _aed(running),
            ]
        )

    title = f"Supplier Statement - {_supplier_name(supplier)}"

    return {
        "title": title,
        "headers": [
            "Date",
            "Type",
            "Reference",
            "Note",
            "Debit",
            "Credit",
            "Balance",
        ],
        "rows": rendered_rows,
    }


# =============================================================================
# ── RENDER: CSV
# =============================================================================

def _render_csv(data: dict) -> HttpResponse:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([data["title"]])
    writer.writerow([f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}"])
    writer.writerow([])
    writer.writerow(data["headers"])
    for row in data["rows"]:
        writer.writerow(row)

    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    safe_name = data["title"].replace(" ", "_").replace("/", "-")
    resp["Content-Disposition"] = f'attachment; filename="{safe_name}.csv"'
    return resp


# =============================================================================
# ── RENDER: XLSX
# =============================================================================

def _render_xlsx(data: dict) -> HttpResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "openpyxl is not installed. Run: pip install openpyxl", status=500
        )

    wb = Workbook()
    ws = wb.active
    ws.title = data["title"][:31]

    headers = data["headers"]
    rows = data["rows"]
    nc = len(headers)

    # ── Styles ────────────────────────────────────────────────────────────────
    thin_side = Side(style="thin", color="CBD5E1")
    bdr = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    hdr_fill = PatternFill(
        start_color="1E3A5F", end_color="1E3A5F", fill_type="solid"
    )
    even_fill = PatternFill(
        start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
    )
    odd_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    total_fill = PatternFill(
        start_color="0F2744", end_color="0F2744", fill_type="solid"
    )
    banner_fill = PatternFill(
        start_color="0F2744", end_color="0F2744", fill_type="solid"
    )

    # ── Row 1: Title banner ───────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    c = ws.cell(row=1, column=1, value=data["title"])
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = banner_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Row 2: Timestamp ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    c = ws.cell(
        row=2,
        column=1,
        value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}",
    )
    c.font = Font(name="Calibri", italic=True, size=9, color="64748B")
    c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Row 3: Spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        c.border = bdr
    ws.row_dimensions[4].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, 5):
        is_total = any(
            str(v).strip().upper().startswith("TOTAL") for v in row
        )
        fill = total_fill if is_total else (even_fill if ri % 2 == 0 else odd_fill)
        padded = list(row) + [""] * max(0, nc - len(row))

        for ci, val in enumerate(padded[:nc], 1):
            c = ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
            c.font = Font(
                name="Calibri",
                size=9,
                bold=is_total,
                color="FFFFFF" if is_total else "0F172A",
            )
            c.fill = fill
            c.border = bdr
            c.alignment = Alignment(
                vertical="center",
                horizontal="right" if ci > 2 else "left",
            )
        ws.row_dimensions[ri].height = 16

    # ── Column widths ─────────────────────────────────────────────────────────
    for col in range(1, nc + 1):
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(4, 5 + len(rows))
            ),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A5"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    safe_name = data["title"].replace(" ", "_").replace("/", "-")
    resp["Content-Disposition"] = f'attachment; filename="{safe_name}.xlsx"'
    return resp


# =============================================================================
# ── RENDER: PDF
# =============================================================================

def _render_pdf(data: dict, fy_id) -> HttpResponse:
    try:
        from reportlab.lib import colors as C
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        return HttpResponse(
            "reportlab is not installed. Run: pip install reportlab", status=500
        )

    headers = data["headers"]
    rows = data["rows"]
    nc = len(headers)

    # ── Page setup ────────────────────────────────────────────────────────────
    use_landscape = nc > 6
    page = landscape(A4) if use_landscape else A4
    PW, PH = page
    MAR = 16 * mm
    CONTENT_W = PW - 2 * MAR

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=MAR,
        rightMargin=MAR,
        topMargin=63 * mm,
        bottomMargin=24 * mm,
        title=data["title"],
    )

    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY = C.HexColor("#0f2744")
    ACCENT = C.HexColor("#1a6bcc")
    SLATE = C.HexColor("#64748b")
    LIGHT_ROW = C.HexColor("#f8fafc")
    GRID_LINE = C.HexColor("#e2e8f0")
    WHITE = C.white
    TOTAL_BG = C.HexColor("#0f2744")
    OUTER_BORDER = C.HexColor("#94a3b8")
    company = _report_company()
    logo_path = _report_logo_path()

    # ── Text styles ───────────────────────────────────────────────────────────
    header_font_size = 7.5 if nc <= 8 else 6.6 if nc <= 10 else 6
    body_font_size = 8 if nc <= 8 else 7.2 if nc <= 10 else 6.6

    ST_HDR = ParagraphStyle(
        "hdr",
        fontName="Helvetica-Bold",
        fontSize=header_font_size,
        textColor=WHITE,
        alignment=TA_CENTER,
        leading=max(header_font_size + 2, 8),
        wordWrap="CJK",
    )
    ST_TDL = ParagraphStyle(
        "tdl",
        fontName="Helvetica",
        fontSize=body_font_size,
        textColor=C.HexColor("#0f172a"),
        alignment=TA_LEFT,
        leading=max(body_font_size + 2, 8),
        wordWrap="CJK",
    )
    ST_TDR = ParagraphStyle(
        "tdr",
        fontName="Helvetica",
        fontSize=body_font_size,
        textColor=C.HexColor("#0f172a"),
        alignment=TA_RIGHT,
        leading=max(body_font_size + 2, 8),
        wordWrap="CJK",
    )
    ST_FTL = ParagraphStyle(
        "ftl",
        fontName="Helvetica-Bold",
        fontSize=body_font_size,
        textColor=WHITE,
        alignment=TA_LEFT,
        leading=max(body_font_size + 2, 8),
        wordWrap="CJK",
    )
    ST_FTR = ParagraphStyle(
        "ftr",
        fontName="Helvetica-Bold",
        fontSize=body_font_size,
        textColor=WHITE,
        alignment=TA_RIGHT,
        leading=max(body_font_size + 2, 8),
        wordWrap="CJK",
    )

    # ── Detect numeric columns (right-align) ──────────────────────────────────
    def _is_numeric_col(ci: int) -> bool:
        for row in rows[:20]:
            if ci < len(row):
                v = str(row[ci])
                if v.startswith("AED") or v.startswith("-AED"):
                    return True
                clean = v.replace(",", "").replace(".", "").replace("-", "").replace("+", "").strip()
                if clean.isdigit() and len(clean) > 0:
                    return True
        return False

    numeric_cols = {i for i in range(nc) if _is_numeric_col(i)}

    # ── FY label ──────────────────────────────────────────────────────────────
    fy_obj = _get_fy(fy_id)
    fy_label = str(fy_obj) if fy_obj else "All Periods"

    # ── Page header/footer callback ───────────────────────────────────────────
    def _on_page(canvas, doc):
        canvas.saveState()

        left = MAR
        right = PW - MAR
        top = PH - MAR
        bottom = MAR - 6 * mm
        width = right - left

        header_h = 30 * mm
        meta_h = 14 * mm
        footer_h = 11 * mm

        canvas.setStrokeColor(OUTER_BORDER)
        canvas.setLineWidth(0.8)
        canvas.rect(left, bottom, width, top - bottom, stroke=1, fill=0)
        canvas.line(left, top - header_h, right, top - header_h)
        canvas.line(left, top - header_h - meta_h, right, top - header_h - meta_h)
        canvas.line(left, bottom + footer_h, right, bottom + footer_h)

        if logo_path:
            canvas.drawImage(
                logo_path,
                left + 8,
                top - 22 * mm,
                width=24 * mm,
                height=16 * mm,
                preserveAspectRatio=True,
                mask="auto",
            )

        company_x = left + 30 * mm
        canvas.setFillColor(C.HexColor("#0f172a"))
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawString(company_x, top - 11 * mm, company["name"][:56])
        canvas.setFillColor(SLATE)
        canvas.setFont("Helvetica", 6.8)
        company_lines = list(company.get("address_lines", []))
        if company.get("city_state_zip"):
            company_lines.append(company["city_state_zip"])
        if company.get("country"):
            company_lines.append(company["country"])
        if company.get("trn"):
            company_lines.append(f"TRN {company['trn']}")
        text_y = top - 15 * mm
        for line in company_lines[:4]:
            canvas.drawString(company_x, text_y, line)
            text_y -= 3.7 * mm

        canvas.setFillColor(NAVY)
        canvas.setFont("Helvetica", 20 if not use_landscape else 18)
        canvas.drawRightString(right - 8, top - 12 * mm, company.get("title", "REPORT"))
        canvas.setFillColor(C.HexColor("#111827"))
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawRightString(right - 8, top - 21 * mm, data["title"][:64])
        canvas.setFillColor(SLATE)
        canvas.setFont("Helvetica", 6.8)
        canvas.drawRightString(
            right - 8,
            top - 25.5 * mm,
            f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}",
        )

        meta_top = top - header_h
        split_x = left + width * 0.48
        canvas.line(split_x, meta_top, split_x, meta_top - meta_h)
        canvas.setFont("Helvetica", 6.8)
        canvas.setFillColor(SLATE)
        left_rows = [
            ("Report Name", data["title"]),
            ("Period", fy_label),
        ]
        right_rows = [
            ("Rows", str(len(rows))),
            ("Page", str(doc.page)),
        ]
        meta_y = meta_top - 5 * mm
        for label, value in left_rows:
            canvas.drawString(left + 6, meta_y, label)
            canvas.drawString(left + 34 * mm, meta_y, f": {value[:42]}")
            meta_y -= 4.4 * mm
        meta_y = meta_top - 5 * mm
        for label, value in right_rows:
            canvas.drawString(split_x + 6, meta_y, label)
            canvas.drawString(split_x + 22 * mm, meta_y, f": {value}")
            meta_y -= 4.4 * mm

        canvas.setFont("Helvetica", 6.8)
        canvas.setFillColor(SLATE)
        canvas.drawString(left + 6, bottom + 3.8 * mm, "Confidential - Internal Use Only")
        canvas.drawCentredString(PW / 2, bottom + 3.8 * mm, data["title"])
        canvas.drawRightString(right - 6, bottom + 3.8 * mm, f"Page {doc.page}")

        canvas.restoreState()

    # ── Build paragraph cells ─────────────────────────────────────────────────
    def _p(val, hdr=False, foot=False, ci=0):
        s = str(val) if val is not None else "—"
        if hdr:
            return Paragraph(s, ST_HDR)
        if foot:
            return Paragraph(s, ST_FTR if (ci in numeric_cols or ci > 0) else ST_FTL)
        return Paragraph(s, ST_TDR if ci in numeric_cols else ST_TDL)

    # ── Column widths: wider for text cols, narrower for numeric ──────────────
    weights = []
    min_weight = 0.8 if nc <= 8 else 0.65
    max_weight = 2.4 if nc <= 8 else 1.9
    sample_rows = rows[: min(len(rows), 20)]
    for ci, header in enumerate(headers):
        max_cell_len = max(
            (len(str(row[ci])) for row in sample_rows if ci < len(row)),
            default=0,
        )
        base_weight = max(len(str(header)), max_cell_len)
        if ci in numeric_cols:
            weight = max(min_weight, min(base_weight / 14, 1.2))
        else:
            weight = max(min_weight, min(base_weight / 10, max_weight))
        weights.append(weight)

    total_weight = sum(weights) or nc
    col_widths = [(CONTENT_W * weight) / total_weight for weight in weights]

    # ── Table data ────────────────────────────────────────────────────────────
    tdata = [[_p(h, hdr=True) for h in headers]]

    for row in rows:
        is_total = any(str(v).strip().upper().startswith("TOTAL") for v in row)
        padded = list(row) + [""] * max(0, nc - len(row))
        tdata.append(
            [_p(padded[ci], foot=is_total, ci=ci) for ci in range(nc)]
        )

    # ── Table style ───────────────────────────────────────────────────────────
    tbl = Table(tdata, colWidths=col_widths, repeatRows=1, splitByRow=True)
    style_cmds = [
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
        ("LINEBELOW",     (0, 0), (-1, 0),  1.2, ACCENT),
        ("TOPPADDING",    (0, 0), (-1, 0),  7),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  7),
        ("LEFTPADDING",   (0, 0), (-1, 0),  5),
        ("RIGHTPADDING",  (0, 0), (-1, 0),  5),
        # Data rows
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_ROW]),
        ("TOPPADDING",    (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING",   (0, 1), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 1), (-1, -1), 5),
        ("LINEBELOW",     (0, 1), (-1, -1), 0.3, GRID_LINE),
        ("BOX",           (0, 0), (-1, -1), 0.5, GRID_LINE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]

    # Highlight TOTAL rows dark
    for ri, row in enumerate(rows, 1):
        if any(str(v).strip().upper().startswith("TOTAL") for v in row):
            style_cmds.append(("BACKGROUND", (0, ri), (-1, ri), TOTAL_BG))
            style_cmds.append(("LINEABOVE",  (0, ri), (-1, ri), 1.0, ACCENT))

    tbl.setStyle(TableStyle(style_cmds))

    # ── Build document ────────────────────────────────────────────────────────
    story = [Spacer(1, 1 * mm), tbl]
    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)

    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    safe_name = data["title"].replace(" ", "_").replace("/", "-")
    resp["Content-Disposition"] = f'attachment; filename="{safe_name}.pdf"'
    return resp
